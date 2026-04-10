"""Excel LCA document parser — extracts structure from spreadsheet-based LCA models.

Parses Excel workbooks containing carbon accounting / LCA calculations and
extracts a structured representation that can be used to generate Mangrove
component YAML. Handles common patterns found in consultancy LCA spreadsheets.
"""

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl


@dataclass
class DataPoint:
    """A single data point extracted from the spreadsheet."""
    name: str
    value: float | str | None
    unit: str
    cell_ref: str  # e.g., "Sheet1!B5"
    is_input: bool = True  # True = input, False = calculated
    is_static: bool = False  # True = emission factor / constant
    category: str = ""
    formula: str | None = None


@dataclass
class Calculation:
    """A calculation step extracted from a formula."""
    name: str
    operator: str  # summation, product, quotient, difference, or expression
    output_ref: str  # cell reference of the output
    output_unit: str
    input_refs: list[str] = field(default_factory=list)
    expression: str | None = None


@dataclass
class ParsedLCA:
    """Complete parsed LCA structure from a workbook."""
    source_file: str
    sheets: list[str]
    data_points: list[DataPoint]
    calculations: list[Calculation]
    emission_factors: list[DataPoint]
    activities: list[DataPoint]
    outputs: list[DataPoint]
    metadata: dict = field(default_factory=dict)

    def summary(self) -> str:
        """Human-readable summary of the parsed LCA."""
        lines = [
            f"Parsed LCA: {self.source_file}",
            f"Sheets: {', '.join(self.sheets)}",
            f"Data points: {len(self.data_points)}",
            f"  Activities (inputs): {len(self.activities)}",
            f"  Emission factors: {len(self.emission_factors)}",
            f"  Calculated outputs: {len(self.outputs)}",
            f"Calculations: {len(self.calculations)}",
        ]
        return "\n".join(lines)


# Heuristic patterns for identifying sheet types
EF_SHEET_PATTERNS = re.compile(r"(?i)(emission.?factor|ef\b|factor|constant|static|param)")
INPUT_SHEET_PATTERNS = re.compile(r"(?i)(input|activit|data|event|collect|monitor)")
CALC_SHEET_PATTERNS = re.compile(r"(?i)(calc|model|emission|result|summar|output|lca)")

# Patterns for detecting emission factor values in cell labels
EF_LABEL_PATTERNS = re.compile(
    r"(?i)(emission.?factor|ef\b|gwp|conversion|density|heating.?value|"
    r"calorific|carbon.?content|moisture|fraction|coefficient)"
)

# Unit patterns
# For matching units inside extracted substrings (parenthesized text, etc.)
_UNIT_WORDS = (
    r"kgCO[₂2]e|tCO[₂2]e|gCO[₂2]e|kg|tonnes?|tons?|gallons?|"
    r"kWh|MWh|MJ|GJ|BTU|miles?|km|liters?|therms?|%"
)
UNIT_PATTERNS = re.compile(rf"(?i)({_UNIT_WORDS})")
# For detecting units in full label text (needs word boundaries to avoid false positives)
UNIT_IN_LABEL = re.compile(rf"(?i)\b({_UNIT_WORDS})\b")


def parse_excel(path: str | Path) -> ParsedLCA:
    """Parse an Excel workbook and extract LCA structure.

    Args:
        path: Path to .xlsx file

    Returns:
        ParsedLCA with extracted data points, calculations, and structure
    """
    path = Path(path)
    wb = openpyxl.load_workbook(str(path), data_only=False)

    all_data_points: list[DataPoint] = []
    all_calculations: list[Calculation] = []

    # First pass: classify sheets
    sheet_roles: dict[str, str] = {}
    for name in wb.sheetnames:
        if EF_SHEET_PATTERNS.search(name):
            sheet_roles[name] = "ef"
        elif INPUT_SHEET_PATTERNS.search(name):
            sheet_roles[name] = "input"
        elif CALC_SHEET_PATTERNS.search(name):
            sheet_roles[name] = "calc"
        else:
            sheet_roles[name] = "unknown"

    # Second pass: extract data from each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        role = sheet_roles.get(sheet_name, "unknown")

        dps, calcs = _extract_sheet(ws, sheet_name, role)
        all_data_points.extend(dps)
        all_calculations.extend(calcs)

    # Classify data points
    emission_factors = [dp for dp in all_data_points if dp.is_static]
    activities = [dp for dp in all_data_points if dp.is_input and not dp.is_static]
    outputs = [dp for dp in all_data_points if not dp.is_input]

    return ParsedLCA(
        source_file=str(path),
        sheets=wb.sheetnames,
        data_points=all_data_points,
        calculations=all_calculations,
        emission_factors=emission_factors,
        activities=activities,
        outputs=outputs,
        metadata={
            "sheet_roles": sheet_roles,
        },
    )


def _extract_sheet(ws, sheet_name: str, role: str) -> tuple[list[DataPoint], list[Calculation]]:
    """Extract data points and calculations from a single worksheet."""
    data_points: list[DataPoint] = []
    calculations: list[Calculation] = []

    # Scan for label-value pairs (label in column A/B, value in adjacent column)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=min(ws.max_column or 1, 20)):
        for i, cell in enumerate(row):
            if cell.value is None:
                continue

            # Look for label cells (strings) with adjacent value cells
            if isinstance(cell.value, str) and cell.value.strip():
                label = cell.value.strip()

                # Skip header-looking cells
                if len(label) < 3 or (label.isupper() and len(label) > 50):
                    continue

                # Look for a value in the next cell(s)
                value_cell = None
                for offset in range(1, min(4, len(row) - i)):
                    candidate = row[i + offset] if i + offset < len(row) else None
                    if candidate and candidate.value is not None:
                        value_cell = candidate
                        break

                if value_cell is None:
                    continue

                cell_ref = f"{sheet_name}!{value_cell.coordinate}"
                unit = _detect_unit(label)
                is_formula = isinstance(value_cell.value, str) and value_cell.value.startswith("=")

                if is_formula:
                    # This is a calculation
                    formula = value_cell.value
                    calc = _parse_formula(label, formula, cell_ref, unit, sheet_name)
                    if calc:
                        calculations.append(calc)
                    data_points.append(DataPoint(
                        name=label,
                        value=None,
                        unit=unit,
                        cell_ref=cell_ref,
                        is_input=False,
                        formula=formula,
                    ))
                else:
                    # This is an input value
                    is_ef = (role == "ef" or bool(EF_LABEL_PATTERNS.search(label)))
                    try:
                        value = float(value_cell.value)
                    except (TypeError, ValueError):
                        value = value_cell.value

                    data_points.append(DataPoint(
                        name=label,
                        value=value,
                        unit=unit,
                        cell_ref=cell_ref,
                        is_input=True,
                        is_static=is_ef,
                        category="emission_factor" if is_ef else "activity",
                    ))

    return data_points, calculations


def _detect_unit(label: str) -> str:
    """Try to detect a unit from a label string."""
    # Check for unit in parentheses: "Mass (tonnes)"
    paren_match = re.search(r"\(([^)]+)\)", label)
    if paren_match:
        unit_candidate = paren_match.group(1).strip()
        if UNIT_PATTERNS.search(unit_candidate):
            return unit_candidate

    # Check for unit after slash or colon
    for pattern in [r"(?:in|per|/)\s*(\S+)", r":\s*(\S+)$"]:
        match = re.search(pattern, label)
        if match and UNIT_PATTERNS.search(match.group(1)):
            return match.group(1)

    # Check the label itself for unit mentions (stricter matching)
    unit_match = UNIT_IN_LABEL.search(label)
    if unit_match:
        return unit_match.group(1)

    return ""


def _parse_formula(name: str, formula: str, cell_ref: str, unit: str, sheet_name: str) -> Calculation | None:
    """Parse an Excel formula into a Calculation."""
    if not formula or not formula.startswith("="):
        return None

    expr = formula[1:]  # Strip leading =

    # Detect operator from formula pattern
    if expr.upper().startswith("SUM(") or expr.upper().startswith("SUMPRODUCT("):
        operator = "summation"
    elif "+" in expr and "*" not in expr and "/" not in expr:
        operator = "summation"
    elif "*" in expr and "+" not in expr and "-" not in expr:
        operator = "product"
    elif "/" in expr and "+" not in expr and "-" not in expr:
        operator = "quotient"
    elif "-" in expr and "+" not in expr and "*" not in expr:
        operator = "difference"
    else:
        operator = "expression"

    # Extract cell references from the formula
    cell_refs = re.findall(r"(?:'?([^'!]+)'?!)?([A-Z]+\d+)", expr)
    input_refs = []
    for ref_sheet, ref_cell in cell_refs:
        ref_sheet = ref_sheet or sheet_name
        input_refs.append(f"{ref_sheet}!{ref_cell}")

    return Calculation(
        name=name,
        operator=operator,
        output_ref=cell_ref,
        output_unit=unit,
        input_refs=input_refs,
        expression=expr,
    )


def to_component_sketch(parsed: ParsedLCA) -> dict:
    """Convert a parsed LCA into a rough component sketch (dict).

    This produces a starting point for component YAML generation.
    The sketch needs human review and refinement, but captures the
    calculation structure from the spreadsheet.
    """
    inputs_section: dict = {
        "event_data_points": [],
        "static_data_points": [],
    }

    for dp in parsed.activities:
        slug = _slugify(dp.name)
        inputs_section["event_data_points"].append({
            "slug_template": slug,
            "unit": dp.unit or "unknown",
            "description": dp.name,
        })

    for dp in parsed.emission_factors:
        slug = _slugify(dp.name)
        inputs_section["static_data_points"].append({
            "slug_template": slug,
            "unit": dp.unit or "unknown",
            "description": dp.name,
        })

    outputs_section: dict = {
        "calculated_data_points": [],
    }
    for dp in parsed.outputs:
        slug = "calculated-" + _slugify(dp.name)
        outputs_section["calculated_data_points"].append({
            "slug_template": slug,
            "unit": dp.unit or "tCO₂e",
            "description": dp.name,
        })

    # Ensure at least one output
    if not outputs_section["calculated_data_points"]:
        outputs_section["calculated_data_points"].append({
            "slug_template": "calculated-total",
            "unit": "tCO₂e",
            "description": "Total calculated output",
        })

    return {
        "component": {
            "id": _slugify(Path(parsed.source_file).stem),
            "name": Path(parsed.source_file).stem.replace("-", " ").title(),
            "version": "1.0",
            "metadata": {
                "pathway": "any",
                "stage": "any",
                "description": f"Auto-generated from {Path(parsed.source_file).name}",
            },
            "inputs": inputs_section,
            "outputs": outputs_section,
            "parameters": [],
            "node_tree": _build_node_tree(parsed),
        }
    }


def _build_node_tree(parsed: ParsedLCA) -> dict:
    """Build a rough node tree from parsed calculations."""
    if not parsed.calculations:
        # No formulas found — create a simple summation of all inputs
        children = []
        for dp in parsed.activities:
            children.append({
                "name": dp.name,
                "data_point_type": _slugify(dp.name),
                "output_unit": dp.unit or "unknown",
            })
        if not children:
            children.append({"name": "Placeholder", "constant": 0})
        return {
            "nexus_nodes_attributes": [{
                "name": "Total",
                "operator": "summation",
                "output_unit": "tCO₂e",
                "nexus_nodes_attributes": children,
            }]
        }

    # Build nodes from calculations
    nodes = []
    for calc in parsed.calculations:
        node: dict = {
            "name": calc.name,
            "operator": calc.operator,
            "output_unit": calc.output_unit or "tCO₂e",
            "data_point_type": "calculated-" + _slugify(calc.name),
        }
        if calc.operator == "expression" and calc.expression:
            node["operator"] = calc.expression

        # Add placeholder children for input refs
        children = []
        for ref in calc.input_refs:
            children.append({
                "name": f"Input from {ref}",
                "output_unit": "",
                "data_point_type": f"TODO-{_slugify(ref)}",
            })
        if children:
            node["nexus_nodes_attributes"] = children

        nodes.append(node)

    return {"nexus_nodes_attributes": nodes}


def _slugify(text: str) -> str:
    """Convert text to a slug (lowercase, hyphens, no special chars)."""
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"[\s_]+", "-", text)
    text = re.sub(r"-+", "-", text)
    text = text.strip("-")
    return text or "unnamed"
