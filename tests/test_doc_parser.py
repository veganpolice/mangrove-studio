"""Tests for the Excel LCA document parser."""

import tempfile
from pathlib import Path

import openpyxl
import pytest
import yaml

from mangrove_studio.agent.doc_parser import (
    DataPoint,
    ParsedLCA,
    parse_excel,
    to_component_sketch,
    _slugify,
    _detect_unit,
)


def _create_simple_lca_workbook(path: str):
    """Create a simple LCA workbook for testing."""
    wb = openpyxl.Workbook()

    # Emission Factors sheet
    ws_ef = wb.active
    ws_ef.title = "Emission Factors"
    ws_ef["A1"] = "Emission Factor"
    ws_ef["B1"] = "Value"
    ws_ef["C1"] = "Unit"
    ws_ef["A2"] = "EF - Electricity (kgCO2e/kWh)"
    ws_ef["B2"] = 0.5
    ws_ef["A3"] = "EF - Diesel (kgCO2e/gallon)"
    ws_ef["B3"] = 10.21
    ws_ef["A4"] = "GWP CH4 to CO2e"
    ws_ef["B4"] = 28.0

    # Activity Data sheet
    ws_input = wb.create_sheet("Activity Data")
    ws_input["A1"] = "Activity"
    ws_input["B1"] = "Quantity"
    ws_input["A2"] = "Electricity use (kWh)"
    ws_input["B2"] = 1500.0
    ws_input["A3"] = "Diesel fuel (gallon)"
    ws_input["B3"] = 200.0
    ws_input["A4"] = "Transport distance (mile)"
    ws_input["B4"] = 350.0

    # Calculations sheet
    ws_calc = wb.create_sheet("Calculations")
    ws_calc["A1"] = "Calculation"
    ws_calc["B1"] = "Result"
    ws_calc["A2"] = "Electricity emissions (tCO2e)"
    ws_calc["B2"] = "='Activity Data'!B2*'Emission Factors'!B2"
    ws_calc["A3"] = "Diesel emissions (tCO2e)"
    ws_calc["B3"] = "='Activity Data'!B3*'Emission Factors'!B3"
    ws_calc["A4"] = "Total emissions (tCO2e)"
    ws_calc["B4"] = "=SUM(B2:B3)"

    wb.save(path)


def _create_biochar_lca_workbook(path: str):
    """Create a more realistic biochar LCA workbook."""
    wb = openpyxl.Workbook()

    # Parameters
    ws = wb.active
    ws.title = "Parameters"
    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    ws["A2"] = "Biochar carbon content fraction"
    ws["B2"] = 0.82
    ws["A3"] = "Moisture content (%)"
    ws["B3"] = 0.15
    ws["A4"] = "Heating value (BTU/lb)"
    ws["B4"] = 12500

    # Production Data
    ws2 = wb.create_sheet("Production Inputs")
    ws2["A1"] = "Input"
    ws2["B1"] = "Value"
    ws2["A2"] = "Biochar produced dry (tonne)"
    ws2["B2"] = 90.0
    ws2["A3"] = "Electricity consumed (kWh)"
    ws2["B3"] = 5000
    ws2["A4"] = "Natural gas used (therms)"
    ws2["B4"] = 150

    # Emission Factors
    ws3 = wb.create_sheet("EF Library")
    ws3["A1"] = "Factor"
    ws3["B1"] = "Value"
    ws3["A2"] = "EF grid electricity (kgCO2e/kWh)"
    ws3["B2"] = 0.42
    ws3["A3"] = "EF natural gas (kgCO2e/therm)"
    ws3["B3"] = 5.3
    ws3["A4"] = "CO2/C conversion factor"
    ws3["B4"] = 3.667

    # Model
    ws4 = wb.create_sheet("LCA Model")
    ws4["A1"] = "Calculation"
    ws4["B1"] = "Result"
    ws4["A2"] = "Gross carbon stored (tCO2e)"
    ws4["B2"] = "='Production Inputs'!B2*'Parameters'!B2*'EF Library'!B4"
    ws4["A3"] = "Electricity emissions (tCO2e)"
    ws4["B3"] = "='Production Inputs'!B3*'EF Library'!B2"
    ws4["A4"] = "Gas emissions (tCO2e)"
    ws4["B4"] = "='Production Inputs'!B4*'EF Library'!B3"
    ws4["A5"] = "Total process emissions (tCO2e)"
    ws4["B5"] = "=B3+B4"
    ws4["A6"] = "Net removal (tCO2e)"
    ws4["B6"] = "=B2-B5"

    wb.save(path)


class TestSlugify:
    def test_basic(self):
        assert _slugify("Electricity Use") == "electricity-use"

    def test_special_chars(self):
        assert _slugify("EF - Diesel (kgCO2e/gallon)") == "ef-diesel-kgco2egallon"

    def test_empty(self):
        assert _slugify("") == "unnamed"


class TestDetectUnit:
    def test_unit_in_parens(self):
        assert _detect_unit("Mass (tonnes)") == "tonnes"

    def test_unit_in_label(self):
        assert _detect_unit("Electricity use kWh") == "kWh"

    def test_no_unit(self):
        assert _detect_unit("Total value") == ""

    def test_percentage(self):
        assert _detect_unit("Moisture content (%)") == "%"


class TestParseExcel:
    def test_parse_simple_workbook(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        assert len(parsed.sheets) == 3
        assert parsed.emission_factors  # Should find EFs
        assert parsed.activities  # Should find activities
        assert parsed.outputs  # Should find calculations
        assert parsed.calculations  # Should find formulas

    def test_sheet_role_detection(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        roles = parsed.metadata["sheet_roles"]
        assert roles["Emission Factors"] == "ef"
        assert roles["Activity Data"] == "input"
        assert roles["Calculations"] == "calc"

    def test_emission_factors_detected(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        ef_names = [dp.name for dp in parsed.emission_factors]
        assert any("Electricity" in name for name in ef_names)
        assert any("Diesel" in name for name in ef_names)

    def test_calculations_detected(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        calc_names = [c.name for c in parsed.calculations]
        assert any("emissions" in name.lower() for name in calc_names)

    def test_summary(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        summary = parsed.summary()
        assert "Data points:" in summary
        assert "Calculations:" in summary


class TestParseBiocharWorkbook:
    def test_parse_biochar_lca(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_biochar_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        assert len(parsed.sheets) == 4
        assert parsed.emission_factors
        assert parsed.calculations

    def test_formula_operator_detection(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_biochar_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        operators = {c.name: c.operator for c in parsed.calculations}
        # Products should be detected (mass * carbon_content * factor)
        assert any(op in ("product", "expression") for op in operators.values())
        # Summation should be detected
        assert any(op == "summation" for op in operators.values())


class TestToComponentSketch:
    def test_sketch_from_simple_workbook(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        sketch = to_component_sketch(parsed)
        comp = sketch["component"]

        assert "id" in comp
        assert "name" in comp
        assert "inputs" in comp
        assert "outputs" in comp
        assert "node_tree" in comp
        assert comp["outputs"]["calculated_data_points"]

    def test_sketch_has_emission_factors_as_static(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        sketch = to_component_sketch(parsed)
        static_dpts = sketch["component"]["inputs"]["static_data_points"]
        assert len(static_dpts) > 0

    def test_sketch_yaml_serializable(self):
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            _create_simple_lca_workbook(f.name)
            parsed = parse_excel(f.name)

        sketch = to_component_sketch(parsed)
        # Should be serializable to YAML without errors
        yaml_str = yaml.dump(sketch, default_flow_style=False, allow_unicode=True)
        assert "component:" in yaml_str
